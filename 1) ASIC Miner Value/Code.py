"""
I need the entire data on page https://www.asicminervalue.com/ --Model, Release, Hashrate, Power, Noise, Algo, Profitability, and the link associated with each model

Using the link associated with each model -you will need to parse the specific page for each of the hardware - For example, for Jasminer - we need all its data -- profitability, Algorithms (Algorythm Hashrate Consumption Efficiency Profitability), Description, All fields in specifications, minable coins, mining pools, all the information in Where to buy and cloud mining information (including the html links).
"""


# IMPORTS:

from requests import get as get_request
from bs4 import BeautifulSoup
from unicodedata import normalize
from os.path import exists
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import startfile


# CONSTANTS:

WEBSITE = 'https://www.asicminervalue.com'
EXCEL_FILE = 'Scraped Data.xlsx'
COLUMN_NAMES = ['#', 'Model', 'Release', 'Hashrate', 'Power', 'Noise', 'Algorithms', 'Profitability', 'Link', 'Description', 'Specifications', 'Minable coins', 'Mining pools', 'Where to buy?', 'Cloud mining']


# CONNECTING TO EXCEL SHEET:

sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name

if not exists(EXCEL_FILE):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
else:
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.create_sheet(title=sheet_title)
    wb.active = sheet

for column_num, column_name in enumerate(iterable=COLUMN_NAMES, start=1):
    sheet.cell(row=1, column=column_num, value=column_name)  # inserting column names

wb.save(EXCEL_FILE)

# print(sheet); exit()  # debugging


# SCRAPING PART 1: ['Model', 'Release', 'Hashrate', 'Power', 'Noise', 'Algo', 'Profitability', 'Link']

website_html_parsed = BeautifulSoup(markup=get_request(url=WEBSITE).text, features='html.parser')
# print(website_html_parsed.prettify()); exit()  # debugging

model_row_html_list = website_html_parsed.tbody.find_all(name='tr')  # list of HTML of all the miner models
# print(model_row_html_list); exit()  # debugging

for sr_num, model_row_html in enumerate(iterable=model_row_html_list, start=1):

    row_num = sr_num + 2

    print(f'\n#{sr_num}')
    sheet.cell(row=row_num, column=1, value=sr_num)

    # print(model_row_html.prettify(), '\n')  # debugging

    # 'Model', 'Release', 'Hashrate', 'Power', 'Noise', 'Algo', 'Profitability':

    div_tags = model_row_html.find_all(name='div')
    # print(div_tags)  # debugging
    values = list(map(lambda div: normalize('NFKD', div.text).strip(), div_tags))  # https://stackoverflow.com/a/34669482
    # print(values)  # debugging
    values = values[3:][:7]  # for a reason 🤫
    # print(values)  # debugging

    # Algos' Name Correction: (if there will be multiple algos of a model, extraction using ".text" will give the number of algos and not their names)
    try:  # checking if the value is a number or not
        int(values[5])
    except ValueError:  # value is not a number => everything's good
        pass
    else:  # value is a number => there were multiple algos of this model
        values[5] = div_tags[8].span['title']  # correcting

    for column_index, value in enumerate(iterable=values, start=1):
        print(f'{COLUMN_NAMES[column_index]}: {value}')
        sheet.cell(row=row_num, column=column_index+1, value=value)

    # 'Link':
    model_page = WEBSITE + model_row_html.a['href']
    print(f'{COLUMN_NAMES[8]}: {model_page}')
    sheet.cell(row=row_num, column=9, value=model_page)

    # SCRAPING PART 2: ['Description', Profitability, Algorithms, 'Specifications', 'Minable coins', 'Mining pools', 'Where to buy?', 'Cloud mining']

    model_page_html_parsed = BeautifulSoup(markup=get_request(url=model_page).text, features='html.parser')
    # print(model_page_html_parsed.prettify()); exit()  # debugging

    container_div = model_page_html_parsed.body.find(name='div', class_='container', recursive=False)
    # print(container_div.prettify()); exit()  # debugging

    # 'Description':
    desc = container_div.p.text
    # print(desc)  # debugging
    desc = ' '.join(desc.split())
    print(f'{COLUMN_NAMES[9]}: {desc}')
    sheet.cell(row=row_num, column=10, value=desc)

    # 'Specifications':
    specs_dict = {}
    for spec_row_html in container_div.find(name='div', class_='col-sm-8').table.find_all(name='tr'):
        specs_dict[spec_row_html.th.text] = spec_row_html.td.text
    print(f'{COLUMN_NAMES[10]}: {specs_dict}')
    sheet.cell(row=row_num, column=11, value=str(specs_dict))

    # 'Minable coins', 'Mining pools', 'Where to buy?', 'Cloud mining':

    for div_tag in container_div.find_all(name='div', class_='col-sm-12'):
        if div_tag.h2 is not None:
            match COLUMN_NAMES.index(div_tag.h2.text):

                case 11:
                    coin_names = []
                    for img_tag in div_tag.find_all(name='img'):
                        coin_name_html = BeautifulSoup(markup=img_tag['title'], features='html.parser')
                        # print(coin_name_html.prettify())  # debugging
                        coin_names.append(coin_name_html.text.removesuffix(coin_name_html.i.text))
                    print(f'{COLUMN_NAMES[11]}: {coin_names}')
                    sheet.cell(row=row_num, column=12, value=', '.join(coin_names))

                case 12:
                    pools_links = [pool_row_html.a['href'] for pool_row_html in div_tag.find_all(name='tr')]
                    print(f'{COLUMN_NAMES[12]}: {pools_links}')
                    sheet.cell(row=row_num, column=13, value=', '.join(pools_links))

                case 13:
                    stores_dict = {}
                    for store_row_html in div_tag.tbody.find_all(name='tr'):
                        stores_dict[store_row_html.b.text] = store_row_html.find(name='a', class_='btn btn-primary')['href']
                    print(f'{COLUMN_NAMES[13]}: {stores_dict}')
                    sheet.cell(row=row_num, column=14, value=str(stores_dict))

                case 14:
                    fixed = {'Provider': {'BitFuFu (Cloud Mining)': 'https://www.bitfufu.com/list?source=086&utm_source=cps&utm_medium=partner&utm_campaign=asicminervalue'}, 'Mining plans': ['Bitcoin (BTC)', 'Ethereum (ETH)', 'BitcoinCash (BCH)', 'Dash (DASH)'], 'Hardware': 'Bitmain (Innosilicon)'}
                    print(f'{COLUMN_NAMES[14]}: {fixed}')
                    sheet.cell(row=row_num, column=15, value=str(fixed))

    wb.save(EXCEL_FILE)  # (after every insertion)


startfile(EXCEL_FILE)  # automatically open Excel Sheet when process completes

print('\nSUCCESS')
