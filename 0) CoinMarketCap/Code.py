# IMPORTS:

from requests import get as get_request
from bs4 import BeautifulSoup
from itertools import count
from json import loads
from os.path import exists
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import startfile


# CONSTANTS:

WEBSITE = 'https://coinmarketcap.com'
EXCEL_FILE = 'Scraped Data.xlsx'


# CONNECTING TO EXCEL SHEET:

sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name

if not exists(EXCEL_FILE):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
else:
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.create_sheet(title=sheet_title)
    wb.active = sheet

# Inserting column names:
for i, column in enumerate(iterable=['#', 'Currency', 'Website', 'Technical Document', 'Explorer', 'Source Code', 'Message Board', 'Chat', 'Announcement', 'Reddit', 'Facebook', 'Twitter'], start=1):
    sheet.cell(row=1, column=i, value=column)

row_num, sr_num = 3, 1

wb.save(EXCEL_FILE)

# print(sheet); exit()  # debugging


# MAIN:

for page_num in count(start=1):  # infinite loop

    webpage = f'{WEBSITE}/?page={page_num}'
    print()  # spacing
    print('=>', webpage, '\n')

    response = get_request(url=webpage)
    # print(response)  # debugging

    if response.status_code == 200:  # everything's okay

        # SCRAPING PART 1:

        soup = BeautifulSoup(markup=response.text, features='html.parser')  # https://www.crummy.com/software/BeautifulSoup/bs4/doc/#differences-between-parsers; https://www.crummy.com/software/BeautifulSoup/bs4/doc/#installing-a-parser
        # print(soup.prettify()); input()  # debugging

        currency_divs = soup.find(name='tbody').find_all(class_=['sc-16r8icm-0 escjiH', 'sc-1rqmhtg-0 jUUSMS'])
        for num, currency_div in enumerate(iterable=currency_divs, start=1):

            num += (page_num-1) * 100

            # print(currency_div.prettify(), '\n'); input()  # debugging

            currency_sub_url = currency_div.a['href']
            # print(f'{num}) {currency_sub_url}'); input()  # debugging

            currency_site = WEBSITE + currency_sub_url
            print(f'{num}) {currency_site} \n')

            # SCRAPING PART 2:

            while True:  # "soup.find(name='script', id='__NEXT_DATA__')" MAY return None, which WILL BE FIXED after some tries
                soup = BeautifulSoup(markup=get_request(url=currency_site).text, features='html.parser')
                # print(soup.prettify()); input()  # debugging

                try:
                    json_str = soup.find(name='script', id='__NEXT_DATA__').text  # all thanks to "View page source"!
                    # print(json_str); input()  # debugging
                    break
                except AttributeError:
                    pass  # try again

            json_dict = loads(s=json_str)  # parse
            # from json import dumps; print(dumps(obj=json_dict, indent=8)); input()  # debugging

            currency_name = json_dict['props']['initialProps']['pageProps']['info']['name']
            print(currency_name)

            urls_dict = json_dict['props']['initialProps']['pageProps']['info']['urls']
            # print(urls_dict); input()  # debugging

            sheet.cell(row=row_num, column=1, value=sr_num)
            sheet.cell(row=row_num, column=2, value=currency_name)

            for i, (attr, url_list) in enumerate(iterable=urls_dict.items(), start=3):

                print(f'{attr}: {url_list}')

                sheet.cell(row=row_num, column=i, value='; '.join(url_list))

            print()  # spacing

            wb.save(EXCEL_FILE)  # (after every insertion)

            row_num += 1
            sr_num += 1

    elif response.status_code == 404:  # webpage not found
        print('Stopping... \n')
        break  # stop looping

    else:
        from time import sleep
        sleep(1)
        exit(response.reason)


startfile(EXCEL_FILE)  # automatically open Excel Sheet when process completes

print('SUCCESS')
