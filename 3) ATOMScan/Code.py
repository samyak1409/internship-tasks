"""
Write a Python script to scrape from https://atomscan.com/blocks/1 till the latest block.
We need to scrape every piece of information there.

We will have to write one more follow-on code that will parse through all these transaction hashes and collect the detailed transactions' data.
See this: https://atomscan.com/transactions/DC6DFDF7C91B14F7FCB6227B487F16F1AA747A8EA17EC8F2294C835C28A1BA4C
"""


# IMPORTS:

from bs4 import BeautifulSoup
from selenium.webdriver import Chrome
from selenium.common.exceptions import WebDriverException
from os.path import exists
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import startfile
from time import sleep


# CONSTANTS:

WEBSITE = 'https://atomscan.com'
EXCEL_FILE = 'Scraped Data.xlsx'
COLUMN_NAMES = ['#', 'URL', 'Time', 'Height', 'Number of Transactions', 'Block Hash', 'Proposer', 'Transaction Hashes', 'Transaction Details']
START = 1  # range: [1, ~10M]


# FUNCTIONS:

def get_parsed_page_html():
    return BeautifulSoup(markup=driver.page_source, features='html.parser')


def wait_to_load():
    print('SLOW INTERNET: PLEASE WAIT...')
    sleep(1)


# CONNECTING TO EXCEL SHEET:

sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name

if not exists(EXCEL_FILE):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
else:
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.create_sheet(title=sheet_title)
    wb.active = sheet

for column_num, column_name in enumerate(iterable=COLUMN_NAMES, start=1):
    sheet.cell(row=1, column=column_num, value=column_name)  # inserting column names

wb.save(EXCEL_FILE)

# print(sheet); exit()  # debugging


# MAIN:

# Website is Dynamic 🤦‍ https://www.geeksforgeeks.org/scrape-content-from-dynamic-websites; https://chromedriver.chromium.org/getting-started

try:
    driver = Chrome()  # webdriver init
except WebDriverException:
    raise SystemExit('''\nERROR: 'chromedriver' executable needs to be in PATH. Please see https://chromedriver.chromium.org/getting-started#h.p_ID_36 \n
TL;DR:
1) Download the ChromeDriver binary for your platform from https://chromedriver.chromium.org/downloads
2) Include the ChromeDriver location in your PATH environment variable''')

driver.minimize_window()  # to show the following info

print('\nMAKE SURE YOU HAVE FAST INTERNET AND LAG-FREE SYSTEM!')

try:
    blocks = int(input('\nHOW MANY?: '))
except ValueError:
    raise SystemExit('\nERROR: EMPTY OR NON-NUMERIC INPUT')
else:
    driver.maximize_window()  # window must not be minimized, else page will load in greater time

    reached_max_height = False

    for height in range(START, START+blocks):

        row_num = height-START + 3

        print(f'\nBlock #{height}')
        sheet.cell(row=row_num, column=1, value=height)

        url = f'{WEBSITE}/blocks/{height}'
        print('URL:', url)
        sheet.cell(row=row_num, column=2, value=url)

        driver.get(url=url)  # open the webpage

        # 'Time', 'Height', 'Number of Transactions', 'Block Hash', 'Proposer':

        already_tried = False  # if 'Proposer' will not load correctly, will try one more time

        while True:  # sometimes block details may take time to load

            try:
                block_details_div = get_parsed_page_html().find(name='div', class_='card-content block-details').div
            except AttributeError:  # "Block Not Found"
                reached_max_height = True
                print('\nSTOPPING: REACHED MAXIMUM BLOCK HEIGHT...')
                break
            # print(block_details_div.prettify()); exit()  # debugging

            block_details = []
            for row_div in block_details_div.find_all(name='div', class_='columns', recursive=False):
                # print(row_div.prettify()); continue  # debugging
                key_div = row_div.div
                value_div = key_div.next_sibling
                block_detail = value_div.text
                if key_div.text == 'Proposer':
                    block_detail = {block_detail: WEBSITE+value_div.a['href']}
                block_details.append(block_detail)

            if not block_details:  # if block_details is empty
                wait_to_load()  # waiting for 'Block Details' to load
                continue

            proposer, validator_link = list(block_details[-1].items())[0]
            if (proposer == validator_link.split('/')[-1]) and (not already_tried):
                wait_to_load()  # waiting for 'Proposer' to load in case still loading
                already_tried = True
                continue  # trying one more time

            break  # while True

        if reached_max_height:
            break

        for column_index, block_detail in enumerate(iterable=block_details, start=2):
            print(f'{COLUMN_NAMES[column_index]}: {block_detail}')
            sheet.cell(row=row_num, column=column_index+1, value=str(block_detail))

        if int(block_details[2]) > 0:  # if transaction(s) exist
            # 'Transaction Hashes':
            while True:  # when there will be more no. of transactions, may not load immediately, will be loaded after a few tries
                hash_dict = {}
                try_again = False
                for txn_row in get_parsed_page_html().tbody.find_all(name='tr'):
                    hash_ = txn_row.td.next_sibling
                    try:
                        hash_dict[hash_.div.text] = WEBSITE + hash_.a['href']
                    except TypeError:  # ('NoneType' object is not subscriptable)
                        try_again = True
                        break  # for loop
                if try_again:
                    wait_to_load()  # waiting for 'Transactions' to load
                    continue
                print('Transaction Hashes:', hash_dict)
                sheet.cell(row=row_num, column=8, value=str(hash_dict))
                break

            # 'Transaction Details':
            txn_details = {}
            for txn_hash, txn_link in hash_dict.items():
                driver.get(url=txn_link)  # open the webpage
                txn_detail = {}
                while True:  # sometimes transaction detail may take time to load
                    for detail_row in get_parsed_page_html().find(name='div', class_='card-content').div.find_all(name='div', class_='columns'):
                        txn_detail[detail_row.div.text] = detail_row.div.next_sibling.text.strip()
                    if not txn_detail:  # if txn_detail is empty
                        wait_to_load()  # waiting for 'Transaction Details' to load
                        continue
                    break
                txn_details[txn_hash] = txn_detail
            print('Transaction Details:', txn_details)
            sheet.cell(row=row_num, column=9, value=str(txn_details))

        wb.save(EXCEL_FILE)  # (after every row insertion)

    startfile(EXCEL_FILE)  # automatically open Excel Sheet when process completes
    print('\nSUCCESS!')

finally:
    driver.quit()
