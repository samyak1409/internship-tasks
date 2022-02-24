# IMPORTS:

from requests import get as get_request, RequestException
from ast import literal_eval
from datetime import datetime, date, timedelta, MINYEAR, MAXYEAR
from os.path import exists
from openpyxl import Workbook, load_workbook


# ATTRIBUTES:

COINS = {'Namecoin': 'nmc',
         'Emercoin': 'emc',
         'Siacoin': 'sc',
         'Gamecredits': 'game',
         'Peercoin': 'ppc',
         'Feathercoin': 'ftc',
         'Stratis': 'strat',
         'Cosmos': 'atom'}

TYPES = {'Tweets Cnt': 'daily_tweets_cnt',
         'Active Address': 'daily_active_address',
         'New Address': 'daily_new_address',
         'Block Cnt': 'daily_block_cnt',
         'Avg Difficulty': 'daily_avg_difficulty',
         'Avg Hashrate': 'daily_avg_hashrate',
         'Fee': 'daily_fee',
         'Fee Usd': 'daily_fee_usd',
         'Avg Block Fee': 'daily_avg_block_fee',
         'Median Block Fee': 'daily_median_block_fee',
         'Avg Tx Fee': 'daily_avg_tx_fee',
         'Avg Tx Fee Usd': 'daily_avg_tx_fee_usd',
         'Sent Value': 'daily_sent_value',
         'Sent Value Usd': 'daily_sent_value_usd',
         'Median Block Sent Value': 'daily_median_block_sent_value',
         'Avg Tx Value': 'daily_avg_tx_value',
         'Tx Cnt': 'daily_tx_cnt',
         'Reward': 'daily_reward',
         'Reward Usd': 'daily_reward_usd',
         'Mining Profitability': 'daily_mining_profitability',
         'Total Mined Coin': 'daily_total_mined_coin',
         'Price': 'daily_price',
         'Marketcap': 'daily_marketcap',
         'Gas Used': 'daily_gas_used',
         'Avg Gas Price': 'daily_avg_gas_price',
         'Avg Gas Limit': 'daily_avg_gas_limit',
         'Avg Tx Gas Used': 'daily_avg_tx_gas_used',
         'Sent Value Gas': 'daily_sent_value_gas',
         'Uncle Block Cnt': 'daily_uncle_block_cnt',
         'Uncle Reward': 'daily_uncle_reward',
         'Uncle Reward Usd': 'daily_uncle_reward_usd',
         'Uncle Total Mined Coin': 'daily_uncle_total_mined_coin'}

OUTPUT_DIR = 'Scraped Data'


# MAIN:

for coin_name, coin_code in COINS.items():

    print()  # spacing
    print(coin_name)

    # EXCEL INIT:
    excel_file = f'{OUTPUT_DIR}\\{coin_name}.xlsx'
    sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name
    if not exists(excel_file):
        wb = Workbook()  # init excel
        sheet = wb.active  # get active sheet
        sheet.title = sheet_title  # rename sheet title
    else:
        wb = load_workbook(excel_file)  # load excel
        sheet = wb.create_sheet(title=sheet_title)  # new sheet with title
        wb.active = sheet  # mark sheet active
    for column_num, column_name in enumerate(iterable=('DATE', *TYPES.keys()), start=1):
        sheet.cell(row=1, column=column_num, value=column_name)  # inserting column names
    wb.save(excel_file)
    # print(sheet); break  # debugging

    # DATA COLLECTION:
    data_dict_list = []
    min_date, max_date = date(year=MAXYEAR, month=12, day=31), date(year=MINYEAR, month=1, day=1)  # bounds

    for column_num, (type_name, type_val) in enumerate(iterable=TYPES.items(), start=2):  # "start=2" coz column 1 is 'DATE'

        link = f'https://{coin_code}.tokenview.com/v2api/chart/?coin={coin_code}&type={type_val}'

        # Getting Request's Response:
        while True:
            try:
                response = get_request(url=link, stream=False, timeout=1)  # stream and timeout parameter -> important
            except RequestException as e:
                print(f'{type(e).__name__}:', e.__doc__.split('\n')[0], 'TRYING AGAIN...')
            else:
                if response.status_code == 200:
                    break
                else:  # bad response
                    print(f'{response.status_code}: {response.reason} TRYING AGAIN...')

        json_dict = response.json()  # (no need of "json" package)

        if json_dict['code'] == 200:  # if data of this particular type of quantity of this coin is there

            data = json_dict['data']  # type = str
            # print(type(data)); exit()  # debugging
            # https://stackoverflow.com/questions/1894269/how-to-convert-string-representation-of-list-to-a-list
            data = literal_eval(node_or_string=data)  # type = list ✔
            # print(type(data)); exit()  # debugging

            if data:  # (data can be empty)

                print(f'{type_name}: {len(data)} ({link})')

                unit = json_dict['unit']
                if unit:  # some data can have a unit
                    sheet.cell(row=1, column=column_num).value += f' ({unit})'
                    # print(sheet.cell(row=1, column=column_num).value)  # debugging

                start_date = date.fromisoformat(data[0].copy().popitem()[0])  # "popitem" remove and return a (key, value) pair as a 2-tuple
                if start_date < min_date:
                    min_date = start_date
                try:
                    last_date = date.fromisoformat(data[-1].copy().popitem()[0])  # date_str to date_obj
                except ValueError:  # sometimes LAST date = 'lastdate'
                    last_date = date.fromisoformat(data[-2].copy().popitem()[0])
                if last_date > max_date:
                    max_date = last_date
                data_dict = {start_date: [mapping.popitem()[1] for mapping in data]}
                # print(data_dict); exit()  # debugging
                data_dict_list.append(data_dict)

    # print(len(data_dict_list), min_date, max_date); break  # debugging

    # DATA OUTPUT:
    date_ = min_date
    for row_num in range(3, (max_date-min_date).days+3+1):  # dates' insertion
        sheet.cell(row=row_num, column=1, value=date_)
        date_ += timedelta(days=1)  # 👌
    for column_num, data_dict in enumerate(iterable=data_dict_list, start=2):
        start_date, values = data_dict.popitem()
        start_row = 0
        for row_num in range(3, sheet.max_row+1):  # finding start date row number
            if sheet.cell(row=row_num, column=1).value == start_date:
                start_row = row_num
                break
        for row_num, value in enumerate(iterable=values, start=start_row):  # data insertion
            sheet.cell(row=row_num, column=column_num, value=value)

    # SAVE EXCEL:
    wb.save(excel_file)

print('\nSUCCESS!')
