"""
https://www.gridcoinstats.eu/block/viewDates
The data to scrape is: All the data rows for columns: [Date, Blocks, Difficulty, Transactions, Receivers, Minted, Fees] for all the pages (~108).
"""


# IMPORTS:

from requests import Session
from bs4 import BeautifulSoup
from os.path import exists
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import startfile


# CONSTANTS:

BASE_URL = 'https://www.gridcoinstats.eu/block/viewDates'
COLUMNS = ['Date (Time (UTC))', 'Blocks (Height)', 'Difficulty (Avg, Min, Max)', 'Transactions (Volume inc. Staked Coins)', 'Receivers (Volume Sent Coins)', 'Minted', 'Fees']
EXCEL = 'Scraped Data.xlsx'


# CONNECTING TO EXCEL FILE:

sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name

if not exists(EXCEL):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
else:
    wb = load_workbook(EXCEL)
    sheet = wb.create_sheet(title=sheet_title)
    wb.active = sheet

sheet.append(COLUMNS)  # INSERTING COLUMN NAMES

wb.save(EXCEL)
# startfile(EXCEL); print(sheet); exit()  # debugging


# REQUESTS SESSION INIT https://docs.python-requests.org/en/latest/user/advanced/#session-objects:

with Session() as session:  # https://en.wikipedia.org/wiki/HTTP_persistent_connection#Advantages

    session.stream = False  # stream off for all the requests of this session

    # GETTING LAST PAGE NUM:

    print('\nGETTING LAST PAGE NUM...', end=' ')

    response = session.get(url=BASE_URL)
    # print(response)  # debugging

    parsed_html = BeautifulSoup(markup=response.text, features='html.parser')
    # print(parsed_html.prettify())  # debugging

    last_page = int(parsed_html.find(name='select', attrs={'name': 'page'}).find_all(name='option')[-1].text)
    print(last_page)

    # MAIN:

    for page in range(1, last_page+1):

        url = f'{BASE_URL}/{page}'
        print(f'\n{page})', url)

        for table_row in BeautifulSoup(markup=session.get(url=url).text, features='html.parser').find(name='table', id='blocksTable').tbody.find_all(name='tr'):

            data_row = []

            for table_data in table_row.find_all(name='td'):

                # print(table_data.text)  # 👎 joint text of different tags

                value = ', '.join(table_data.strings)  # 👍 https://www.crummy.com/software/BeautifulSoup/bs4/doc/#strings-and-stripped-strings
                # print(value)  # debugging

                data_row.append(value)

            print(data_row)
            sheet.append(data_row)  # https://openpyxl.readthedocs.io/en/stable/#:~:text=Sample%20code

        wb.save(EXCEL)  # (after data insertion of every page)


startfile(EXCEL)  # automatically open Excel Sheet when process completes
print('\nSUCCESS!')
