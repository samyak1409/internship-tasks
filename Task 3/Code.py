"""
Write a Python script to scrape from https://atomscan.com/blocks/1 till https://atomscan.com/blocks/9272083.
We need to scrape every piece of information there.
"""


# IMPORTS:

from bs4 import BeautifulSoup
from selenium.webdriver import Chrome
from selenium.common.exceptions import WebDriverException
from os.path import exists
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import startfile
from time import sleep


# ATTRIBUTES:

WEBSITE = 'https://atomscan.com'
EXCEL_FILE = 'Scraped Data.xlsx'
COLUMN_NAMES = ['#', 'URL', 'Time', 'Height', 'Number of Transactions', 'Block Hash', 'Proposer', 'Transaction Hashes', 'Transaction Details']
MAXIMIZE_CHROME = False
START = 9286724  # range: [1, ~10M]


# FUNCTIONS:

def get_parsed_page_html():
    return BeautifulSoup(markup=driver.page_source, features='html.parser')  # https://www.crummy.com/software/BeautifulSoup/bs4/doc/#differences-between-parsers


def wait_to_load():
    print('PLEASE WAIT...')
    sleep(1)


# CONNECTING TO EXCEL SHEET:

sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name

if not exists(EXCEL_FILE):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
else:
    wb = load_workbook(EXCEL_FILE)
    sheet = wb.create_sheet(title=sheet_title)
    wb.active = sheet

for column_num, column_name in enumerate(iterable=COLUMN_NAMES, start=1):
    sheet.cell(row=1, column=column_num, value=column_name)  # inserting column names

wb.save(EXCEL_FILE)

# print(sheet); exit()  # debugging


# MAIN:

# Website is Dynamic 🤦‍ https://www.geeksforgeeks.org/scrape-content-from-dynamic-websites; https://chromedriver.chromium.org/getting-started

try:
    driver = Chrome()  # webdriver init
except WebDriverException:
    raise SystemExit('''\nERROR: 'chromedriver' executable needs to be in PATH. Please see https://chromedriver.chromium.org/getting-started#h.p_ID_36 \n
TL;DR:
1) Download the ChromeDriver binary for your platform from https://chromedriver.chromium.org/downloads
2) Include the ChromeDriver location in your PATH environment variable''')

driver.minimize_window()  # to show the following info

print('\nMAKE SURE YOU HAVE A FAST INTERNET CONNECTION AND LAG-FREE SYSTEM!')

try:
    blocks = int(input('\nHow many?: '))
except ValueError:
    raise SystemExit('\nERROR: EMPTY OR NON-NUMERIC INPUT')
else:
    if MAXIMIZE_CHROME:
        driver.maximize_window()

    reached_max_height = False

    for height in range(START, START+blocks):

        row_num = height-START + 3

        print(f'\nBlock #{height}')
        sheet.cell(row=row_num, column=1, value=height)

        url = f'{WEBSITE}/blocks/{height}'
        print('URL:', url)
        sheet.cell(row=row_num, column=2, value=url)

        driver.get(url=url)  # open the webpage

        # 'Time', 'Height', 'Number of Transactions', 'Block Hash', 'Proposer':
        for try_ in range(2):  # if 'Proposer' will not load correctly, will try one more time

            try:
                block_details_div = get_parsed_page_html().find(name='div', class_='card-content block-details').div
            except AttributeError:  # "Block Not Found"
                reached_max_height = True
                print('\nSTOPPING: REACHED MAXIMUM BLOCK HEIGHT...')
                break
            # print(block_details_div.prettify()); exit()  # debugging

            values = []
            for row_div in block_details_div.find_all(name='div', class_='columns', recursive=False):
                # print(row_div.prettify()); continue  # debugging
                key_div = row_div.div
                value_div = key_div.next_sibling
                value = value_div.text
                if key_div.text == 'Proposer':
                    value = {value: WEBSITE+value_div.a['href']}
                values.append(value)

            proposer, validator_link = list(values[-1].items())[0]
            if proposer == validator_link.split('/')[-1]:
                if try_ == 0:
                    wait_to_load()  # waiting for 'Proposer' to load in case still loading
            else:
                break

        if reached_max_height:
            break

        for column_index, value in enumerate(iterable=values, start=2):
            print(f'{COLUMN_NAMES[column_index]}: {value}')
            sheet.cell(row=row_num, column=column_index+1, value=str(value))

        if int(values[2]) > 0:  # if transaction(s) exist
            # 'Transaction Hashes':
            while True:  # when there will be more no. of transactions, may not load immediately, will be loaded after a few tries
                hash_dict = {}
                try_again = False
                for txn_row in get_parsed_page_html().tbody.find_all(name='tr'):
                    hash_ = txn_row.td.next_sibling
                    try:
                        hash_dict[hash_.div.text] = WEBSITE + hash_.a['href']
                    except TypeError:  # ('NoneType' object is not subscriptable)
                        try_again = True
                        break  # for loop
                if try_again:
                    wait_to_load()  # waiting for 'Transactions' to load
                    continue
                print('Transaction Hashes:', hash_dict)
                sheet.cell(row=row_num, column=8, value=str(hash_dict))
                break

            # 'Transaction Details':
            txn_details = {}
            for txn_hash, txn_link in hash_dict.items():
                driver.get(url=txn_link)  # open the webpage
                txn_detail = {}
                for detail_row in get_parsed_page_html().find(name='div', class_='card-content').div.find_all(name='div', class_='columns'):
                    txn_detail[detail_row.div.text] = detail_row.div.next_sibling.text.strip()
                txn_details[txn_hash] = txn_detail
            print('Transaction Details:', txn_details)
            sheet.cell(row=row_num, column=9, value=str(txn_details))

        wb.save(EXCEL_FILE)  # (after every row insertion)

    startfile(EXCEL_FILE)  # automatically open Excel Sheet when process completes
    print('\nSUCCESS!')

finally:
    driver.quit()
