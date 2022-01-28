# I need the entire data on page https://www.asicminervalue.com/ --Model, Release, Hashrate, Power, Noise, Algo, Profitability, and the link associated with each model
# Using the link associated with each model -you will need to parse the specific page for each of the hardware - For example, for Jasminer - we need all its data -- profitability, Algorithms (Algorythm Hashrate Consumption Efficiency Profitability), Description, All fields in specifications, minable coins, mining pools, all the information in Where to buy and cloud mining information (including the html links).


# IMPORTS:

from requests import get as get_request
from bs4 import BeautifulSoup
from unicodedata import normalize
from os.path import exists
from openpyxl import Workbook, load_workbook
from datetime import datetime
from os import startfile


# ATTRIBUTES:

website = 'https://www.asicminervalue.com'
excel_file = 'Scraped Data.xlsx'
column_names = ['#', 'Model', 'Release', 'Hashrate', 'Power', 'Noise', 'Algorithms', 'Profitability', 'Link', 'Description', 'Specifications', 'Minable coins', 'Mining pools', 'Where to buy?', 'Cloud mining']


# CONNECTING TO EXCEL SHEET:

sheet_title = str(datetime.now()).replace(':', ';')  # ':' not allowed as an Excel sheet name

if not exists(excel_file):
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_title
else:
    wb = load_workbook(excel_file)
    sheet = wb.create_sheet(title=sheet_title)
    wb.active = sheet

for column_num, column_name in enumerate(iterable=column_names, start=1):
    sheet.cell(row=1, column=column_num, value=column_name)  # inserting column names

wb.save(excel_file)

# print(sheet); exit()  # debugging


# SCRAPING PART 1: ['Model', 'Release', 'Hashrate', 'Power', 'Noise', 'Algo', 'Profitability', 'Link']

website_html_parsed = BeautifulSoup(markup=get_request(url=website).text, features='html.parser')  # https://www.crummy.com/software/BeautifulSoup/bs4/doc/#differences-between-parsers
# print(website_html_parsed.prettify()); exit()  # debugging

model_row_html_list = website_html_parsed.find(name='tbody').find_all(name='tr')  # list of HTML of all the miner models
# print(model_row_html_list); exit()  # debugging

for sr_num, model_row_html in enumerate(iterable=model_row_html_list, start=1):

    row_num = sr_num + 2

    print(f'\n{sr_num}')
    sheet.cell(row=row_num, column=1, value=sr_num)

    # print(model_row_html.prettify(), '\n')  # debugging

    # 'Model', 'Release', 'Hashrate', 'Power', 'Noise', 'Algo', 'Profitability':

    div_tags = model_row_html.find_all(name='div')
    # print(div_tags)  # debugging
    values = list(map(lambda div_tag: normalize('NFKD', div_tag.text).strip(), div_tags))  # https://stackoverflow.com/a/34669482
    # print(values)  # debugging
    values = values[3:][:7]  # for a reason 🤫
    # print(values)  # debugging

    # Algos' Name Correction: (if there will be multiple algos of a model, extraction using ".text" will give the number of algos and not their names)
    try:  # checking if the value is a number or not
        int(values[5])
    except ValueError:  # value is not a number => everything's good
        pass
    else:  # value is a number => there were multiple algos of this model
        values[5] = div_tags[8].span['title']  # correcting

    for column_num, value in enumerate(iterable=values, start=2):
        print(value)
        sheet.cell(row=row_num, column=column_num, value=value)

    # 'Link':
    model_page = website + model_row_html.a['href']
    print(model_page)
    sheet.cell(row=row_num, column=9, value=model_page)

    # SCRAPING PART 2: ['Description', Profitability, Algorithms, 'Specifications', 'Minable coins', 'Mining pools', 'Where to buy?', 'Cloud mining']

    model_page_html_parsed = BeautifulSoup(markup=get_request(url=model_page).text, features='html.parser')
    # print(model_page_html_parsed.prettify()); exit()  # debugging

    container_div = model_page_html_parsed.body.find(name='div', class_='container', recursive=False)
    # print(container_div.prettify()); exit()  # debugging

    # 'Description':
    desc = container_div.p.text
    # print(desc)  # debugging
    desc = ' '.join(desc.split())
    print(desc)
    sheet.cell(row=row_num, column=10, value=desc)

    # 'Specifications':
    specs_dict = {}
    for spec_row_html in container_div.find(name='div', class_='col-sm-8').table.find_all(name='tr'):
        specs_dict[spec_row_html.th.text] = spec_row_html.td.text
    print(specs_dict)
    sheet.cell(row=row_num, column=11, value=str(specs_dict))

    # 'Minable coins', 'Mining pools', 'Where to buy?', 'Cloud mining':

    for div_tag in container_div.find_all(name='div', class_='col-sm-12'):
        if div_tag.h2 is not None:
            match div_tag.h2.text:

                case 'Minable coins':
                    coin_names = []
                    for img_tag in div_tag.find_all(name='img'):
                        x = BeautifulSoup(markup=img_tag['title'], features='html.parser')
                        coin_names.append(x.text.removesuffix(x.i.text))
                    print(coin_names)
                    sheet.cell(row=row_num, column=12, value=', '.join(coin_names))

                case 'Minable pools':
                    pass

                case 'Where to buy?':
                    pass

                case 'Cloud mining':
                    pass

    wb.save(excel_file)  # (after every insertion)


startfile(excel_file)  # automatically open Excel Sheet when process completes

print('\nSUCCESS')
