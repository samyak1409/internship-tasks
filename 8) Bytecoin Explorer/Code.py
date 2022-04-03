"""
Bytecoin: Here, you will need to get the data from the webpage. Make sure to get all the information including the
number of transactions (https://explorer.bytecoin.org/block?height=2033923)
"""


# IMPORTS:

from requests import Session, RequestException
from bs4 import BeautifulSoup
from json import dumps
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook, load_workbook
from os import stat, chdir, startfile, rename
from os.path import exists
from glob import iglob
from time import perf_counter, sleep


start_time = perf_counter()


# CONSTANTS:

BASE_URL = 'https://explorer.bytecoin.org'
HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.51 Safari/537.36'}
COLUMNS = ('Height', 'Version', 'Timestamp', 'Base Reward', 'Transactions Fee', 'Transactions Size', 'Already Generated Transactions', 'Already Generated Key Outputs', 'Transaction Count', 'Transaction Hashes')
DEBUG = False  # (default: False)
THREADS = 1 if DEBUG else 100  # number of concurrent threads to run at once
DATA_DIR = 'Scraped Data'  # path to data dir
MAX_SIZE = 1  # of Excel, you want to be (in MB)


# MAIN FUNCTION:

def main(block_height: int) -> None:

    # Getting Request's Response:
    while True:
        try:
            response = session.get(url=f'{BASE_URL}/block?height={block_height}')
        except RequestException as e:
            print(f'{type(e).__name__}:', e.__doc__.split('\n')[0], 'TRYING AGAIN...')
            sleep(1)  # take a breath
        else:
            if response.status_code == 200:
                break
            else:  # bad response
                print(f'{response.status_code}: {response.reason} TRYING AGAIN...')
                sleep(1)  # take a breath

    soup = BeautifulSoup(markup=response.text, features='html.parser')
    # print(soup.prettify())  # debugging

    block_table, txns_table = soup.find_all(name='table', class_='border_table')[:2]
    # print(block_table)  # debugging
    # print(block_table)  # debugging

    data = {'Height': block_height, 'Version': None, 'Timestamp': None, 'Base Reward': None, 'Transactions Fee': None, 'Transactions Size': None, 'Already Generated Transactions': None, 'Already Generated Key Outputs': None, 'Transaction Count': None, 'Transaction Hashes': None}

    for tr in block_table.find_all(name='tr'):
        key, value = map(lambda td: td.string if td.string else list(td.stripped_strings), tr.find_all(name='td'))
        # print(tr); print(key, value)  # debugging
        # 'Version', 'Timestamp', 'Base Reward', 'Transactions Fee', 'Transactions Size', 'Already Generated Transactions', 'Already Generated Key Outputs':
        match key:
            case 'Version':
                data[key] = float(value[0])
            case 'Timestamp':
                data[key] = value[1]
            case 'Base Reward':
                data[key] = float(value[0])
            case 'Transactions Fee':
                data[key] = float(value[0])
            case 'Transactions Size':
                data[key] = int(value.split(' • ')[0])
            case 'Already Generated Transactions':
                data[key] = int(value)
            case 'Already Generated Key Outputs':
                data[key] = int(value)

    txn_hashes = []
    for tr in txns_table.find_all(name='tr')[1:]:  # skipping the column names row
        txn_hash = next(tr.find(name='td', class_='fixedfont').stripped_strings)
        # print(txn_hash)  # debugging
        txn_hashes.append(txn_hash)
    data['Transaction Count'], data['Transaction Hashes'] = len(txn_hashes), ', '.join(txn_hashes)

    print('\n' + f'{block_height})', 'Transactions:', data['Transaction Count'])

    if DEBUG:
        print(dumps(data, indent=4))

    data_dict[block_height] = list(data.values())


# SESSION INIT:

with Session() as session:

    session.headers = HEADERS
    session.stream = False  # stream off for all the requests of this session

    chdir(DATA_DIR)

    print('\n' + 'Getting start and stop block number...')
    try:
        start = max(map(lambda name: int(name.split('.xlsx')[0]), iglob('*.xlsx')))
    except ValueError:  # no files were there in the dir
        start = 1
    stop = int(next(filter(lambda num: num.isnumeric(), list(BeautifulSoup(session.get(f'{BASE_URL}/block?height=0').text, 'html.parser').a.parent.stripped_strings)[1].split())))
    stop -= stop % 100  # 2541960 -> 2541900
    print(start, stop)

    # Excel Init:
    excel = f'{start}.xlsx'
    if not exists(excel):
        wb = Workbook()
        sheet = wb.active
        sheet.append(COLUMNS)  # writing column names in Excel; https://openpyxl.readthedocs.io/en/stable
        wb.save(excel)
    else:
        wb = load_workbook(excel)
        sheet = wb.active

    # THREADING:

    new = False

    for height in range(start, stop, THREADS):  # start, stop, step

        if stat(excel).st_size >= MAX_SIZE * 1024**2:  # at most {MAX_SIZE} MB of data in one Excel
            wb = Workbook()
            sheet = wb.active
            sheet.append(COLUMNS)  # writing column names in Excel; https://openpyxl.readthedocs.io/en/stable
            new = True

        # Using dict so that the data is kept sorted:
        data_dict = {num: None for num in range(height, height+THREADS)}  # {block_height: block_data}

        # Executing {THREADS} no. of threads at once:
        with ThreadPoolExecutor() as Exec:
            Exec.map(main, range(height, height+THREADS))

        # Writing the data scraped from {THREADS} no. of threads to the Excel:
        for row in filter(None, data_dict.values()):  # filter -> consider only non-empty values
            sheet.append(row)
        # Saving:
        if not new:
            wb.save(excel)  # same excel
            excel_, excel = excel, f'{height+THREADS}.xlsx'
            rename(src=excel_, dst=excel)  # so that if the process stops anytime, will be resumed from there only in future
        else:
            excel = f'{height+THREADS}.xlsx'
            wb.save(excel)  # new excel created
            new = False  # reset

        if DEBUG:
            startfile(excel)
            break


print('\n' + f'Successfully finished in {int(perf_counter()-start_time)}s.')
